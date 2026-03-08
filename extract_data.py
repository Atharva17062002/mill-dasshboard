import openpyxl
import json

wb = openpyxl.load_workbook('/Users/atharvaudavant/development/Mill dashboard/Untitled spreadsheet.xlsx', data_only=True)
ws = wb['Database']

headers = []
for col in range(1, ws.max_column + 1):
    cell = ws.cell(row=1, column=col)
    headers.append(cell.value if cell.value else f"col_{col}")

records = []
for row in range(2, ws.max_row + 1):
    # Check if row has any data
    has_data = False
    for col in range(1, ws.max_column + 1):
        if ws.cell(row=row, column=col).value is not None:
            has_data = True
            break
    if not has_data:
        continue
    
    record = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=row, column=col).value
        if val is not None:
            header = headers[col - 1]
            if hasattr(val, 'isoformat'):
                val = val.isoformat()
            record[header] = val
    if record:
        records.append(record)

print(f"Total records: {len(records)}")
print(json.dumps(records[:3], indent=2, default=str))

# Now compute dashboard metrics
dashboard_ws = wb['Dashboard']
print("\n\nDashboard computed values:")
for row in range(1, 13):
    label = dashboard_ws.cell(row=row, column=1).value
    val = dashboard_ws.cell(row=row, column=2).value
    c_val = dashboard_ws.cell(row=row, column=3).value
    d_val = dashboard_ws.cell(row=row, column=4).value
    e_val = dashboard_ws.cell(row=row, column=5).value
    f_val = dashboard_ws.cell(row=row, column=6).value
    print(f"  Row {row}: {label} = B:{val}, C:{c_val}, D:{d_val}, E:{e_val}, F:{f_val}")

# Save records as JSON
with open('/Users/atharvaudavant/development/Mill dashboard/dashboard/src/data/database.json', 'w') as f:
    json.dump(records, f, indent=2, default=str)
print(f"\nSaved {len(records)} records to database.json")
